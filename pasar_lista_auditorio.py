#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import sys
import os
import time
from datetime import datetime

# Determinar la ruta base:
# Si está "frozen" (ejecutable), se usará el directorio que contiene el exe.
if getattr(sys, 'frozen', False):
    BASE_PATH = os.path.dirname(sys.executable)
else:
    BASE_PATH = os.path.abspath(".")

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout,
    QLabel, QPushButton, QMessageBox, QDateEdit
)
from PyQt6.QtCore import QDate, QThread, pyqtSignal, Qt, QTimer, QSharedMemory
from openpyxl import load_workbook, Workbook
from smartcard.System import readers
from smartcard.CardMonitoring import CardMonitor, CardObserver


##############################################################################
#                           Clase NdefManager (Solo lectura)                 #
##############################################################################

class NdefManager:
    """
    Se encarga de leer un mensaje NDEF (se asume que es una URL con la matrícula)
    a partir de un objeto de conexión de PySCard.
    """
    def __init__(self, connection):
        self.conexion = connection

    def _leer_bloque(self, pagina_inicio):
        if not self.conexion:
            return None

        apdu = [0xFF, 0x00, 0x00, 0x00, 0x04,
                0xD4, 0x42, 0x30, pagina_inicio]
        respuesta, sw1, sw2 = self.conexion.transmit(apdu)
        if sw1 == 0x90 and sw2 == 0x00:
            return respuesta[-16:] if len(respuesta) > 16 else respuesta
        else:
            print(f"Error al leer bloque desde la página {pagina_inicio}: SW1={hex(sw1)} SW2={hex(sw2)}")
            return None

    def leer_ndef(self):
        if not self.conexion:
            return ""

        time.sleep(0.2)
        datos_leidos = []
        current_page = 4
        max_page = 40
        while current_page < max_page:
            bloque = self._leer_bloque(current_page)
            if bloque is None:
                break
            datos_leidos.extend(bloque)
            if 0xFE in bloque:
                break
            current_page += 4

        try:
            indice = datos_leidos.index(0x03)
        except ValueError:
            return ""

        if indice + 1 >= len(datos_leidos):
            return ""

        ndef_len = datos_leidos[indice + 1]
        mensaje_ndef = datos_leidos[indice + 2: indice + 2 + ndef_len]
        if len(mensaje_ndef) < 5:
            return ""

        payload_len = mensaje_ndef[2]
        if len(mensaje_ndef) < 4 + payload_len:
            return ""

        payload = mensaje_ndef[4:4 + payload_len]
        if not payload:
            return ""

        # Quitar prefijo "https://" o "http://", quedándose con la matrícula
        return "".join(chr(b) for b in payload[1:])


##############################################################################
#           Observador de inserción de tarjetas para CardMonitor             #
##############################################################################

class CardReaderObserver(CardObserver):
    def __init__(self, signal_emitter):
        super().__init__()
        self.signal_emitter = signal_emitter

    def update(self, observable, actions):
        added_cards, removed_cards = actions
        if added_cards:
            # Notificar que se inició la lectura (estado "reading": fondo naranja)
            self.signal_emitter.emit_estado("reading", "Leyendo tarjeta...")
            for card in added_cards:
                try:
                    connection = card.createConnection()
                    connection.connect()
                    ndef = NdefManager(connection)
                    matricula = ndef.leer_ndef()
                    if matricula.startswith("https://"):
                        matricula = matricula[len("https://"):]
                    elif matricula.startswith("http://"):
                        matricula = matricula[len("http://"):]
                    if matricula:
                        self.signal_emitter.emit_matricula(matricula)
                    else:
                        self.signal_emitter.emit_estado("error", "Lectura fallida. Pase la tarjeta nuevamente.")
                except Exception as e:
                    # Si se quita muy rápido la tarjeta se lanza este error
                    if "0x80100069" in str(e):
                        print("Error al conectar o leer tarjeta:", e)
                    else:
                        print("Error al conectar o leer tarjeta:", e)
                    self.signal_emitter.emit_estado("error", "Lectura fallida. Pase la tarjeta nuevamente.")
        if removed_cards:
            # Al retirar la tarjeta, se actualiza el estado a "waiting" (fondo rojo)
            self.signal_emitter.emit_estado("waiting", "Esperando lecturas...")


##############################################################################
#                          Hilo para monitorear tarjetas                     #
##############################################################################

class CardMonitorThread(QThread):
    nuevaMatricula = pyqtSignal(str)
    estadoLectura = pyqtSignal(str, str)  # estado, mensaje

    def __init__(self):
        super().__init__()
        self._stop_flag = False

    def run(self):
        r = readers()
        if not r:
            print("No se encontraron lectores de tarjetas.")
            return

        monitor = CardMonitor()
        observer = CardReaderObserver(self)
        monitor.addObserver(observer)

        while not self._stop_flag:
            time.sleep(0.2)

        monitor.deleteObserver(observer)

    def stop(self):
        self._stop_flag = True

    def emit_matricula(self, matricula):
        self.nuevaMatricula.emit(matricula)

    def emit_estado(self, estado, mensaje):
        self.estadoLectura.emit(estado, mensaje)


##############################################################################
#                     Ventana principal PyQt6 (UI y lógica)                  #
##############################################################################

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("Control de Asistencia NFC")
        self.resize(400, 200)

        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout()
        central_widget.setLayout(main_layout)

        # Selector de fecha en formato dd/MM/yyyy (ej.: 21/02/2025)
        self.date_edit = QDateEdit(QDate.currentDate())
        self.date_edit.setDisplayFormat("dd/MM/yyyy")
        main_layout.addWidget(self.date_edit)

        self.btn_pasar_lista = QPushButton("Pasar Lista")
        self.btn_pasar_lista.clicked.connect(self.toggle_pasar_lista)
        main_layout.addWidget(self.btn_pasar_lista)

        self.label_matricula = QLabel("")
        self.label_matricula.setAlignment(Qt.AlignmentFlag.AlignCenter)
        # Configurar un tamaño de fuente mayor para la matrícula
        self.label_matricula.setStyleSheet("font-size: 28pt; background-color: #FFCDD2;")
        main_layout.addWidget(self.label_matricula)

        self.lectura_en_proceso = False
        self.thread_lectura = None

        # Variables para almacenar las matrículas leídas (sin repeticiones) y las no válidas
        self.matriculas_registradas = set()
        self.invalid_matriculas = set()
        self.registro_matriculas = set()  # Matrículas válidas registradas en el Excel

        # La ruta del Excel se obtiene a partir de BASE_PATH (la carpeta que contiene el exe)
        self.RUTA_EXCEL = os.path.join(BASE_PATH, "asistencia.xlsx")

    def toggle_pasar_lista(self):
        if not self.lectura_en_proceso:
            # Verificar que se pueda abrir el Excel (no esté en uso o se cree si no existe)
            try:
                wb = load_workbook(self.RUTA_EXCEL)
                wb.close()
            except PermissionError:
                QMessageBox.critical(self, "Archivo en uso", "Cierra el archivo de Excel para pasar lista.")
                return
            except FileNotFoundError:
                QMessageBox.warning(self, "Archivo no encontrado",
                                    "No se encontró el archivo asistencia.xlsx. Se creará uno nuevo.")
                wb_new = Workbook()
                wb_new.active.cell(row=1, column=1, value="MATRICULA")
                wb_new.save(self.RUTA_EXCEL)
                wb_new.close()

            # Cargar las matrículas válidas desde el Excel para validación inmediata
            try:
                wb = load_workbook(self.RUTA_EXCEL)
                hoja = wb.active
                self.registro_matriculas = set()
                for row in range(2, hoja.max_row + 1):
                    valor = hoja.cell(row=row, column=1).value
                    if valor is not None:
                        self.registro_matriculas.add(str(valor).strip())
                wb.close()
            except Exception as e:
                QMessageBox.critical(self, "Error", f"No se pudo abrir el archivo de Excel: {e}")
                return

            # Reiniciar los conjuntos para la sesión
            self.matriculas_registradas = set()
            self.invalid_matriculas = set()

            self.lectura_en_proceso = True
            self.btn_pasar_lista.setText("Detener")
            self.thread_lectura = CardMonitorThread()
            self.thread_lectura.nuevaMatricula.connect(self.procesar_matricula)
            self.thread_lectura.estadoLectura.connect(self.actualizar_estado)
            self.thread_lectura.start()
            # Estado inicial: esperando lectura (fondo rojo)
            self.actualizar_estado("waiting", "Esperando lecturas...")
        else:
            self.detener_lectura()

    def detener_lectura(self):
        self.lectura_en_proceso = False
        self.btn_pasar_lista.setText("Pasar Lista")
        if self.thread_lectura is not None:
            self.thread_lectura.stop()
            self.thread_lectura.wait()
            self.thread_lectura = None
        # Al detener, se guarda en el Excel de una sola vez
        self.guardar_asistencia()

    def procesar_matricula(self, matricula):
        # Validar la matrícula contra el registro cargado desde Excel
        if matricula in self.registro_matriculas:
            if matricula not in self.matriculas_registradas:
                self.matriculas_registradas.add(matricula)
                self.actualizar_estado("success", f"Matrícula {matricula} registrada")
            else:
                # Ya se había registrado previamente
                self.actualizar_estado("duplicate", f"Matrícula {matricula} ya fue registrada")
        else:
            if matricula not in self.invalid_matriculas:
                self.invalid_matriculas.add(matricula)
                self.actualizar_estado("error", f"Matrícula {matricula} no está registrada")
                print(f"Matrícula {matricula} no está registrada en el Excel.")

    def actualizar_estado(self, estado, mensaje):
        """
        Actualiza la interfaz según el estado de la lectura.
        Estados:
         - reading: fondo naranja
         - success: fondo verde
         - waiting: fondo rojo
         - error: fondo amarillo
         - duplicate: fondo negro con letras blancas
        Se utiliza un QTimer para volver al estado "waiting" tras 1.5 segundos en success, error o duplicate.
        """
        if estado == "reading":
            style = "background-color: #FFCC80; font-size: 28pt; color: black;"
        elif estado == "success":
            style = "background-color: #C8E6C9; font-size: 28pt; color: black;"
        elif estado == "waiting":
            style = "background-color: #FFCDD2; font-size: 28pt; color: black;"
        elif estado == "error":
            style = "background-color: #FFF9C4; font-size: 28pt; color: black;"
        elif estado == "duplicate":
            style = "background-color: #000000; font-size: 28pt; color: #FFFFFF;"
        else:
            style = "background-color: #FFFFFF; font-size: 28pt; color: black;"

        self.label_matricula.setStyleSheet(style)
        self.label_matricula.setText(mensaje)

        # Para estados que indiquen finalización de la lectura (success, error, duplicate)
        if estado in ["success", "error", "duplicate"]:
            QTimer.singleShot(1500, lambda: self.label_matricula.setStyleSheet("background-color: #FFCDD2; font-size: 28pt; color: black;"))
            QTimer.singleShot(1500, lambda: self.label_matricula.setText("Esperando lecturas..."))

    def guardar_asistencia(self):
        """Actualiza el archivo Excel con las asistencias acumuladas en la sesión."""
        try:
            wb = load_workbook(self.RUTA_EXCEL)
        except Exception:
            QMessageBox.critical(self, "Error", "No se pudo abrir el archivo de Excel para guardar la asistencia.")
            return

        hoja = wb.active
        fecha_seleccionada = self.date_edit.date().toString("dd/MM/yyyy")
        col_fecha = self._find_or_create_date_column(hoja, fecha_seleccionada)

        # Recorrer todas las filas y marcar asistencia según si la matrícula fue leída
        for row in range(2, hoja.max_row + 1):
            celda_matricula = hoja.cell(row=row, column=1).value
            if celda_matricula is not None:
                matricula = str(celda_matricula).strip()
                if matricula in self.matriculas_registradas:
                    hoja.cell(row=row, column=col_fecha, value="SI")
                else:
                    hoja.cell(row=row, column=col_fecha, value="NO")
            else:
                hoja.cell(row=row, column=col_fecha, value="NO")

        try:
            wb.save(self.RUTA_EXCEL)
        except PermissionError:
            QMessageBox.critical(self, "Archivo en uso", "Cierra el archivo de Excel para guardar la asistencia.")
        finally:
            wb.close()
        self.label_matricula.setText("Asistencia guardada.")

    def _get_last_date_column(self, sheet):
        last_valid = 1
        for col in range(2, sheet.max_column + 1):
            val = sheet.cell(row=1, column=col).value
            if val is not None and str(val).strip() != "":
                last_valid = col
        return last_valid

    def _find_or_create_date_column(self, sheet, date_str):
        last_valid = self._get_last_date_column(sheet)
        for col in range(2, last_valid + 1):
            cell_value = sheet.cell(row=1, column=col).value
            if cell_value is None:
                continue
            if isinstance(cell_value, datetime):
                cell_str = cell_value.strftime("%d/%m/%Y")
            else:
                cell_str = str(cell_value).strip()
            if cell_str == date_str:
                return col
        new_col = last_valid + 1
        sheet.cell(row=1, column=new_col, value=date_str)
        return new_col

    def _find_column(self, sheet, date_str):
        last_valid = self._get_last_date_column(sheet)
        for col in range(2, last_valid + 1):
            cell_value = sheet.cell(row=1, column=col).value
            if cell_value is None:
                continue
            if isinstance(cell_value, datetime):
                cell_str = cell_value.strftime("%d/%m/%Y")
            else:
                cell_str = str(cell_value).strip()
            if cell_str == date_str:
                return col
        return None


def main():
    app = QApplication(sys.argv)
    # Evitar múltiples instancias usando QSharedMemory
    shared_memory = QSharedMemory("ControlAsistenciaNFCUniqueKey")
    if not shared_memory.create(1):
        QMessageBox.critical(None, "Instancia ya en ejecución", "El programa ya está en ejecución.")
        sys.exit(0)

    ventana = MainWindow()
    ventana.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()